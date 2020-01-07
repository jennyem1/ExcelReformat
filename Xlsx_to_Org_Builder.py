import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
import xlsxwriter
from xlsxwriter import Workbook
import itertools
from itertools import *

 
#path to the source file
input_file = "C:\\Users\\jen080519\\OneDrive - Paycor, Inc\\Documents\\XLS2ORG_Build\\SourceOrgData_117617.xlsx"

#opening the source file
wb = openpyxl.load_workbook(input_file)
ws = wb.active

#These next loops go through each column and puts the data into arrays to store code values.
code1_list = []
for cell in ws['A']:
    if cell.value is not None:
        if cell.value != 'Level 1 Code':
            code1_list.append(str(cell.value))
    else:
        continue

desc1_list = []
for cell in ws['B']:
    if cell.value is not None:
        if cell.value != 'Level 1 Description':
            desc1_list.append(cell.value)
    else:
        continue

code2_list = []
for cell in ws['C']:
    if cell.value is not None:
        if cell.value != 'Level 2 Code':
            code2_list.append(str(cell.value))
    else:
        continue

desc2_list = []
for cell in ws['D']:
    if cell.value is not None:
        if cell.value != 'Level 2 Description':
            desc2_list.append(cell.value)
    else:
        continue

code3_list = []
for cell in ws['E']:
    if cell.value is not None:
        if cell.value != 'Level 3 Code':
            code3_list.append(str(cell.value))
    else:
        continue

desc3_list = []
for cell in ws['F']:
    if cell.value is not None:
        if cell.value != 'Level 3 Description':
            desc3_list.append(cell.value)
    else:
        continue

code4_list = []
for cell in ws['G']:
    if cell.value is not None:
        if cell.value != 'Level 4 Code':
            code4_list.append(str(cell.value))
    else:
        continue

desc4_list = []
for cell in ws['H']:
    if cell.value is not None:
        if cell.value != 'Level 4 Description':
            desc4_list.append(cell.value)
    else:
        continue

code5_list = []
for cell in ws['I']:
    if cell.value is not None:
        if cell.value != 'Level 5 Code':
            code5_list.append(str(cell.value))
    else:
        continue

desc5_list = []
for cell in ws['J']:
    if cell.value is not None:
        if cell.value != 'Level 5 Description':
            desc5_list.append(cell.value)
    else:
        continue

#Getting the client ID from user and assigning it to the clientID variable.
clientID = [input('Please enter the client ID number: ')]

#building lists for parent codes via f-strings, itertools.product and list comprehension(for parentCode2)
#pulling the information from the original code lists above
parentCode1 = ['']
parentCode2 = [str(a) for a in code1_list]
parentCode3 = [f'{a}{b}' for a, b in itertools.product(code1_list, code2_list)]
parentCode4 = [f'{a}{b}{c}' for a, b, c in itertools.product(code1_list, code2_list, code3_list)]
parentCode5 = [f'{a}{b}{c}{d}' for a, b, c, d in itertools.product(code1_list, code2_list, code3_list, code4_list)]


#We need to find the lenghts of levels 2 - 5 codes in order to know how many times to repeat the data in the column
code1len = len(code1_list)
code2len = len(code2_list)
code3len = len(code3_list)
code4len = len(code4_list)
code5len = len(code5_list)

#Repeating the code levels based on the number of parent levels in each code
#Establishing the length of parent codes needed first
parentCode2_repeat = [parentCode2]*code2len
parentCode2_final = []
for x in parentCode2_repeat:
    for y in x:
        parentCode2_final.append(y)
parentCode2_final.sort()

parentCode3_repeat = [parentCode3]*code3len
parentCode3_final = []
for x in parentCode3_repeat:
    for y in x:
        parentCode3_final.append(y)
parentCode3_final.sort()

parentCode4_repeat = [parentCode4]*code4len
parentCode4_final = []
for x in parentCode4_repeat:
    for y in x:
        parentCode4_final.append(y)
parentCode4_final.sort()

parentCode5_repeat = [parentCode5]*code5len
parentCode5_final = []
for x in parentCode5_repeat:
    for y in x:
        parentCode5_final.append(y)
parentCode5_final.sort()

code2_list_repeat = [code2_list]*code1len
code2_list_final = []
for x in code2_list_repeat:
    for y in x:
        code2_list_final.append(y)

code3_list_repeat = [code3_list]*code2len
code3_list_final = []
for x in code3_list_repeat:
    for y in x:
        code3_list_final.append(y)

code4_list_repeat = [code4_list]*code3len*code2len
code4_list_final = []
for x in code4_list_repeat:
    for y in x:
        code4_list_final.append(y)

code5_list_repeat = [code5_list]*code4len*code3len*code2len
code5_list_final = []
for x in code5_list_repeat:
    for y in x:
        code5_list_final.append(y)

desc3_list_repeat = [desc3_list]*code2len
desc3_list_final = []
for x in desc3_list_repeat:
    for y in x:
        desc3_list_final.append(y)

desc4_list_repeat = [desc4_list]*code3len*code2len
desc4_list_final = []
for x in desc4_list_repeat:
    for y in x:
        desc4_list_final.append(y)

desc5_list_repeat = [desc5_list]*code4len*code3len*code2len
desc5_list_final = []
for x in desc5_list_repeat:
    for y in x:
        desc5_list_final.append(y)


#combining the codes and descriptions into lists in order to write the lists to a specific column
code = code1_list + code2_list_final + code3_list_final + code4_list_final + code5_list_final
description = desc1_list + desc2_list + desc3_list_final + desc4_list_final + desc5_list_final
parentCode = parentCode1 + parentCode2_final + parentCode3_final + parentCode4_final + parentCode5_final


#This adds the user-input client ID to fill the lenght of Column A depending up on the length of the entire worksheet.
totalColsLen = len(code)
clientID_final = [item for item in clientID for x in range(totalColsLen)]

#This adds the action code "a" for the action code column since this is an org build, the code will always be "a", this script is not used for edits
actionCode = 'a'
actionCode_final = [item for item in actionCode for x in range(totalColsLen)]

#using xlsxwriter, create a new workbook/sheet
org_wb = xlsxwriter.Workbook('C:\\Users\\jen080519\\OneDrive - Paycor, Inc\\Documents\\XLS2ORG_Build\\Org_to_Perform.xlsx')
org_ws = org_wb.add_worksheet()

#Adding formatting to the column headers
cell_format = org_wb.add_format()
cell_format2 = org_wb.add_format()
cell_format.set_bold(True)
cell_format.set_font_color('red')
cell_format2.set_bold(True)

#setting up headers
org_ws.write('A1', 'clientId', cell_format)
org_ws.write('B1', 'code', cell_format)
org_ws.write('C1', 'newCode', cell_format2)
org_ws.write('D1', 'description', cell_format)
org_ws.write('E1', 'parentCode', cell_format)
org_ws.write('F1', 'action', cell_format2)

#append list values to excel spreadsheet using xlsxwriter
#need a way to write clientID for all populated cells in columns
org_ws.write_column('A2', clientID_final)
org_ws.write_column('B2', code)
org_ws.write_column('D2', description)
org_ws.write_column('E2', parentCode)
org_ws.write_column('F2', actionCode_final)

org_wb.close()


